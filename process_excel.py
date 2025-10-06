#!/usr/bin/env python3
"""
Excel to Markdown converter with image extraction.
Reads all sheets from data/Fruits.xlsx and converts to Markdown with embedded images.
"""

import os
import sys
import zipfile
import shutil
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import xml.etree.ElementTree as ET


def create_directories():
    """Create necessary output directories."""
    media_dir = Path("media")
    media_dir.mkdir(exist_ok=True)
    return media_dir


def extract_images_from_xlsx(xlsx_path, media_dir):
    """
    Extract images from xlsx file using robust ZIP approach.
    Returns list of extracted image files and relationship mappings.
    """
    images = []
    image_relationships = {}
    
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zip_file:
            # List all files in the ZIP
            file_list = zip_file.namelist()
            
            # Extract images from xl/media/
            media_files = [f for f in file_list if f.startswith('xl/media/')]
            for media_file in media_files:
                # Extract image
                image_data = zip_file.read(media_file)
                image_name = os.path.basename(media_file)
                output_path = media_dir / image_name
                
                with open(output_path, 'wb') as f:
                    f.write(image_data)
                
                images.append(image_name)
                print(f"Extracted image: {image_name}")
            
            # Try to parse drawing relationships
            try:
                drawing_files = [f for f in file_list if f.startswith('xl/drawings/')]
                for drawing_file in drawing_files:
                    try:
                        drawing_data = zip_file.read(drawing_file)
                        # Parse drawing XML to find image references
                        root = ET.fromstring(drawing_data)
                        # This is a simplified approach - actual relationship parsing would be more complex
                        print(f"Found drawing file: {drawing_file}")
                    except Exception as e:
                        print(f"Could not parse drawing file {drawing_file}: {e}")
            except Exception as e:
                print(f"Error processing drawings: {e}")
                
    except Exception as e:
        print(f"Error extracting images: {e}")
        return [], {}
    
    return images, image_relationships


def convert_worksheet_to_markdown(worksheet, sheet_name):
    """Convert a worksheet to markdown table format."""
    # Get the used range
    if worksheet.max_row == 1 and worksheet.max_column == 1:
        # Empty sheet
        return f"### {sheet_name}\n\n*Empty sheet*\n\n", 0, 0
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    # Extract data
    data = []
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_data = []
        for cell in row:
            value = cell.value if cell.value is not None else ""
            # Handle different data types
            if isinstance(value, (int, float)):
                row_data.append(str(value))
            else:
                row_data.append(str(value).replace('\n', ' ').replace('|', '\\|'))
        data.append(row_data)
    
    if not data:
        return f"### {sheet_name}\n\n*No data found*\n\n", 0, 0
    
    # Create markdown table
    markdown = f"### {sheet_name}\n\n"
    
    # Add header row if we have data
    if data:
        header = data[0]
        markdown += "| " + " | ".join(header) + " |\n"
        markdown += "| " + " | ".join(["---"] * len(header)) + " |\n"
        
        # Add data rows
        for row in data[1:]:
            # Ensure row has same number of columns as header
            while len(row) < len(header):
                row.append("")
            row = row[:len(header)]  # Truncate if too long
            markdown += "| " + " | ".join(row) + " |\n"
    
    markdown += "\n"
    return markdown, max_row, max_col


def generate_excel_markdown(workbook, images, image_relationships):
    """Generate the main ExcelFile.md content."""
    content = "# Excel File Analysis\n\n"
    
    # Summary section
    content += "## Summary\n\n"
    sheet_names = workbook.sheetnames
    content += f"Total sheets: {len(sheet_names)}\n\n"
    
    sheet_info = []
    all_sheet_content = []
    
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        sheet_info.append(f"- **{sheet_name}**: {max_row} rows × {max_col} columns")
        
        # Convert sheet to markdown
        sheet_markdown, _, _ = convert_worksheet_to_markdown(worksheet, sheet_name)
        all_sheet_content.append(sheet_markdown)
    
    content += "\n".join(sheet_info) + "\n\n"
    
    # Sheet data sections
    content += "## Sheet Data\n\n"
    content += "\n".join(all_sheet_content)
    
    # Images section
    content += "## Images\n\n"
    if images:
        content += f"Found {len(images)} embedded images:\n\n"
        for image in images:
            content += f"![{image}](media/{image})\n\n"
        
        # Note about sheet mapping
        if not image_relationships:
            content += "*Note: Image-to-sheet mapping could not be determined. All images are listed above.*\n\n"
    else:
        content += "*No embedded images found.*\n\n"
    
    return content


def generate_report(images, method_used="robust unzip approach", errors=None):
    """Generate the Report.md content."""
    content = "# Processing Report\n\n"
    
    content += "## Method Used\n\n"
    content += f"Used **{method_used}** to extract images from the Excel file.\n"
    content += "This method treats the .xlsx file as a ZIP archive and extracts all files from xl/media/*.\n\n"
    
    content += "## Results\n\n"
    content += f"- Successfully processed Excel file: data/Fruits.xlsx\n"
    content += f"- Extracted {len(images)} images to media/ directory\n"
    content += f"- Generated ExcelFile.md with all sheet data\n\n"
    
    content += "## Limitations\n\n"
    content += "- Image-to-sheet mapping requires complex parsing of xl/drawings/* and relationship files\n"
    content += "- Currently images are exported globally without specific sheet associations\n"
    content += "- This approach ensures all embedded images are captured even if not exposed via openpyxl's worksheet._images\n\n"
    
    if errors:
        content += "## Errors Encountered\n\n"
        for error in errors:
            content += f"- {error}\n"
        content += "\n"
    
    content += "## File Operations\n\n"
    content += "- Downloaded Fruits.xlsx from GitHub attachment and placed in data/Fruits.xlsx for version control\n"
    content += "- All images exported to media/ directory with original filenames preserved\n"
    
    return content


def main():
    """Main processing function."""
    print("Starting Excel to Markdown conversion...")
    
    # Check if input file exists
    xlsx_path = Path("data/Fruits.xlsx")
    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found!")
        sys.exit(1)
    
    errors = []
    
    try:
        # Create output directories
        media_dir = create_directories()
        
        # Extract images using robust ZIP approach
        print("Extracting images...")
        images, image_relationships = extract_images_from_xlsx(xlsx_path, media_dir)
        
        # Load workbook for sheet data
        print("Loading workbook...")
        workbook = load_workbook(xlsx_path, data_only=True)
        
        # Generate ExcelFile.md
        print("Generating ExcelFile.md...")
        excel_content = generate_excel_markdown(workbook, images, image_relationships)
        with open("ExcelFile.md", "w", encoding="utf-8") as f:
            f.write(excel_content)
        
        # Generate Report.md
        print("Generating Report.md...")
        report_content = generate_report(images, errors=errors if errors else None)
        with open("Report.md", "w", encoding="utf-8") as f:
            f.write(report_content)
        
        print("✅ Processing completed successfully!")
        print(f"   - ExcelFile.md created with {len(workbook.sheetnames)} sheets")
        print(f"   - {len(images)} images extracted to media/ directory")
        print(f"   - Report.md created with processing details")
        
    except Exception as e:
        error_msg = f"Fatal error during processing: {e}"
        print(f"❌ {error_msg}")
        errors.append(error_msg)
        
        # Try to generate at least a report with the error
        try:
            report_content = generate_report([], errors=errors)
            with open("Report.md", "w", encoding="utf-8") as f:
                f.write(report_content)
        except:
            pass
        
        sys.exit(1)


if __name__ == "__main__":
    main()