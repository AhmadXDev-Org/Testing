#!/usr/bin/env python3
"""
Excel to Markdown converter with robust image extraction.
Reads Excel workbook, converts sheets to Markdown, and extracts all embedded images.
"""

import os
import zipfile
import json
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl import load_workbook
from PIL import Image
import shutil

class ExcelProcessor:
    def __init__(self, excel_path: str, output_dir: str = "."):
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)
        self.media_dir = self.output_dir / "media"
        self.media_dir.mkdir(exist_ok=True)
        
        # Results storage
        self.sheets_data = {}
        self.extracted_images = []
        self.errors = []
        self.processing_log = []
        
    def log(self, message: str):
        """Log processing steps"""
        print(f"[LOG] {message}")
        self.processing_log.append(message)
        
    def process_excel(self):
        """Main processing function"""
        self.log(f"Starting Excel processing for: {self.excel_path}")
        
        try:
            # Step 1: Read sheets and convert to tables
            self.read_excel_sheets()
            
            # Step 2: Extract images using multiple methods
            self.extract_images_via_openpyxl()
            self.extract_images_via_zip()
            
            # Step 3: Generate output files
            self.generate_markdown()
            self.generate_report()
            
        except Exception as e:
            error_msg = f"Critical error during processing: {str(e)}"
            self.log(error_msg)
            self.errors.append(error_msg)
            self.generate_report()
            
    def read_excel_sheets(self):
        """Read all sheets from Excel file and convert to tables"""
        try:
            self.log("Loading Excel workbook...")
            workbook = load_workbook(self.excel_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                self.log(f"Processing sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Get all data from the sheet
                data = []
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                if max_row == 1 and max_col == 1 and worksheet.cell(1, 1).value is None:
                    # Empty sheet
                    self.sheets_data[sheet_name] = {
                        'data': [],
                        'rows': 0,
                        'cols': 0,
                        'is_empty': True
                    }
                    continue
                
                for row in worksheet.iter_rows(min_row=1, max_row=max_row, 
                                               min_col=1, max_col=max_col):
                    row_data = []
                    for cell in row:
                        value = cell.value
                        if value is None:
                            value = ""
                        row_data.append(str(value))
                    data.append(row_data)
                
                self.sheets_data[sheet_name] = {
                    'data': data,
                    'rows': max_row,
                    'cols': max_col,
                    'is_empty': False
                }
                
                self.log(f"Sheet {sheet_name}: {max_row} rows, {max_col} columns")
                
        except Exception as e:
            error_msg = f"Error reading Excel sheets: {str(e)}"
            self.log(error_msg)
            self.errors.append(error_msg)
            
    def extract_images_via_openpyxl(self):
        """Extract images using openpyxl library"""
        try:
            self.log("Attempting image extraction via openpyxl...")
            workbook = load_workbook(self.excel_path)
            
            image_count = 0
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Check for images in the worksheet
                if hasattr(worksheet, '_images') and worksheet._images:
                    self.log(f"Found {len(worksheet._images)} images in sheet: {sheet_name}")
                    
                    for idx, image in enumerate(worksheet._images):
                        try:
                            # Save image
                            image_name = f"openpyxl_{sheet_name}_{idx}.png"
                            image_path = self.media_dir / image_name
                            
                            # Get image data
                            img_data = image._data()
                            with open(image_path, 'wb') as f:
                                f.write(img_data)
                            
                            self.extracted_images.append({
                                'filename': image_name,
                                'path': image_path,
                                'sheet': sheet_name,
                                'method': 'openpyxl',
                                'anchor': getattr(image, 'anchor', 'unknown'),
                                'description': 'To be analyzed'
                            })
                            
                            image_count += 1
                            self.log(f"Extracted image: {image_name}")
                            
                        except Exception as e:
                            error_msg = f"Error extracting image {idx} from sheet {sheet_name}: {str(e)}"
                            self.log(error_msg)
                            self.errors.append(error_msg)
                            
            self.log(f"openpyxl method extracted {image_count} images")
            
        except Exception as e:
            error_msg = f"Error in openpyxl image extraction: {str(e)}"
            self.log(error_msg)
            self.errors.append(error_msg)
            
    def extract_images_via_zip(self):
        """Extract images by treating xlsx as zip file"""
        try:
            self.log("Attempting image extraction via ZIP method...")
            
            with zipfile.ZipFile(self.excel_path, 'r') as zip_file:
                # List all files in the archive
                file_list = zip_file.namelist()
                
                # Find images in xl/media/
                media_files = [f for f in file_list if f.startswith('xl/media/')]
                self.log(f"Found {len(media_files)} files in xl/media/")
                
                for media_file in media_files:
                    try:
                        # Extract the file
                        file_data = zip_file.read(media_file)
                        
                        # Get filename
                        original_name = os.path.basename(media_file)
                        image_name = f"zip_{original_name}"
                        image_path = self.media_dir / image_name
                        
                        # Save the image
                        with open(image_path, 'wb') as f:
                            f.write(file_data)
                        
                        # Try to determine image format
                        try:
                            with Image.open(image_path) as img:
                                img_format = img.format
                                img_size = img.size
                        except:
                            img_format = "Unknown"
                            img_size = (0, 0)
                        
                        self.extracted_images.append({
                            'filename': image_name,
                            'path': image_path,
                            'sheet': 'unknown',  # Will try to map later
                            'method': 'zip',
                            'original_path': media_file,
                            'format': img_format,
                            'size': img_size,
                            'description': 'To be analyzed'
                        })
                        
                        self.log(f"Extracted image via ZIP: {image_name} ({img_format}, {img_size})")
                        
                    except Exception as e:
                        error_msg = f"Error extracting {media_file}: {str(e)}"
                        self.log(error_msg)
                        self.errors.append(error_msg)
                
                # Try to map images to sheets by examining drawings
                self.map_images_to_sheets(zip_file, file_list)
                
        except Exception as e:
            error_msg = f"Error in ZIP image extraction: {str(e)}"
            self.log(error_msg)
            self.errors.append(error_msg)
            
    def map_images_to_sheets(self, zip_file: zipfile.ZipFile, file_list: List[str]):
        """Try to map images to sheets using drawings relationships"""
        try:
            self.log("Attempting to map images to sheets...")
            
            # Look for drawing files
            drawing_files = [f for f in file_list if f.startswith('xl/drawings/')]
            drawing_rels = [f for f in file_list if f.startswith('xl/drawings/_rels/')]
            
            self.log(f"Found {len(drawing_files)} drawing files and {len(drawing_rels)} relationship files")
            
            # This is a complex process that would require parsing XML relationships
            # For now, we'll note the limitation
            if drawing_files or drawing_rels:
                self.log("Drawing files found - image-to-sheet mapping is possible but complex")
                self.log("Will note this limitation in the report")
            
        except Exception as e:
            error_msg = f"Error mapping images to sheets: {str(e)}"
            self.log(error_msg)
            self.errors.append(error_msg)
            
    def generate_markdown(self):
        """Generate the main ExcelFile.md output"""
        try:
            self.log("Generating ExcelFile.md...")
            
            md_content = []
            
            # Header
            md_content.append("# Excel File Analysis: Fruits.xlsx\n")
            
            # Summary section
            md_content.append("## Summary\n")
            md_content.append("| Sheet Name | Rows | Columns | Status |")
            md_content.append("|------------|------|---------|--------|")
            
            for sheet_name, sheet_info in self.sheets_data.items():
                status = "Empty" if sheet_info['is_empty'] else "Data"
                md_content.append(f"| {sheet_name} | {sheet_info['rows']} | {sheet_info['cols']} | {status} |")
            
            md_content.append("")
            
            # Sheet sections
            for sheet_name, sheet_info in self.sheets_data.items():
                md_content.append(f"## Sheet: {sheet_name}\n")
                
                if sheet_info['is_empty']:
                    md_content.append("*This sheet is empty.*\n")
                    continue
                
                # Create markdown table
                data = sheet_info['data']
                if data:
                    # Use first row as headers if it looks like headers
                    headers = data[0]
                    rows = data[1:] if len(data) > 1 else []
                    
                    # Create table header
                    md_content.append("| " + " | ".join(headers) + " |")
                    md_content.append("| " + " | ".join(["---"] * len(headers)) + " |")
                    
                    # Add data rows
                    for row in rows:
                        # Pad row if it's shorter than headers
                        while len(row) < len(headers):
                            row.append("")
                        # Truncate if longer
                        row = row[:len(headers)]
                        md_content.append("| " + " | ".join(row) + " |")
                
                md_content.append("")
            
            # Images section
            md_content.append("## Extracted Images\n")
            
            if not self.extracted_images:
                md_content.append("*No images were found in the Excel file.*\n")
            else:
                md_content.append(f"Found {len(self.extracted_images)} images:\n")
                
                for img_info in self.extracted_images:
                    md_content.append(f"### {img_info['filename']}\n")
                    md_content.append(f"![{img_info['filename']}](media/{img_info['filename']})\n")
                    md_content.append(f"**Method:** {img_info['method']}")
                    md_content.append(f"**Sheet:** {img_info.get('sheet', 'unknown')}")
                    if 'format' in img_info:
                        md_content.append(f"**Format:** {img_info['format']}")
                    if 'size' in img_info:
                        md_content.append(f"**Size:** {img_info['size'][0]}x{img_info['size'][1]}")
                    md_content.append(f"**Description:** {img_info['description']}")
                    md_content.append("")
            
            # Write to file
            output_path = self.output_dir / "ExcelFile.md"
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(md_content))
            
            self.log(f"Generated ExcelFile.md at: {output_path}")
            
        except Exception as e:
            error_msg = f"Error generating markdown: {str(e)}"
            self.log(error_msg)
            self.errors.append(error_msg)
            
    def generate_report(self):
        """Generate the Report.md file"""
        try:
            self.log("Generating Report.md...")
            
            report_content = []
            
            # Header
            report_content.append("# Processing Report: Excel to Markdown Conversion\n")
            
            # File info
            report_content.append("## File Information\n")
            report_content.append(f"**Source File:** {self.excel_path}")
            report_content.append(f"**File Size:** {self.excel_path.stat().st_size} bytes")
            report_content.append(f"**Processing Date:** {Path().cwd()}")
            report_content.append("")
            
            # Methods used
            report_content.append("## Methods Used\n")
            report_content.append("### Sheet Reading")
            report_content.append("- Used `openpyxl.load_workbook()` with `data_only=True`")
            report_content.append("- Extracted all sheets and converted to markdown tables")
            report_content.append("")
            
            report_content.append("### Image Extraction")
            report_content.append("1. **openpyxl method:** Attempted to access `worksheet._images`")
            report_content.append("2. **ZIP method:** Treated .xlsx as ZIP and extracted from `xl/media/`")
            report_content.append("")
            
            # Results
            report_content.append("## Results\n")
            report_content.append(f"**Sheets processed:** {len(self.sheets_data)}")
            report_content.append(f"**Images extracted:** {len(self.extracted_images)}")
            
            if self.extracted_images:
                openpyxl_count = len([img for img in self.extracted_images if img['method'] == 'openpyxl'])
                zip_count = len([img for img in self.extracted_images if img['method'] == 'zip'])
                report_content.append(f"  - Via openpyxl: {openpyxl_count}")
                report_content.append(f"  - Via ZIP method: {zip_count}")
            
            report_content.append("")
            
            # Limitations
            report_content.append("## Limitations and Notes\n")
            report_content.append("- **Image-to-cell mapping:** Complex XML relationship parsing not fully implemented")
            report_content.append("- **Image analysis:** Semantic descriptions require manual browser-based analysis")
            report_content.append("- **Format support:** Focused on standard xlsx format")
            report_content.append("")
            
            # Errors
            if self.errors:
                report_content.append("## Errors Encountered\n")
                for i, error in enumerate(self.errors, 1):
                    report_content.append(f"{i}. {error}")
                report_content.append("")
            
            # Processing log
            report_content.append("## Processing Log\n")
            for log_entry in self.processing_log:
                report_content.append(f"- {log_entry}")
            
            # Write to file
            output_path = self.output_dir / "Report.md"
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(report_content))
            
            self.log(f"Generated Report.md at: {output_path}")
            
        except Exception as e:
            error_msg = f"Error generating report: {str(e)}"
            print(f"[ERROR] {error_msg}")

def main():
    """Main function"""
    excel_path = "data/Fruits.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
    
    processor = ExcelProcessor(excel_path)
    processor.process_excel()
    
    print("\n=== Processing Complete ===")
    print(f"Generated files:")
    print(f"- ExcelFile.md")
    print(f"- Report.md")
    print(f"- Extracted images in media/ directory")

if __name__ == "__main__":
    main()