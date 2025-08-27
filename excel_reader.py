#!/usr/bin/env python3
"""
Excel File Reader Script
Demonstrates reading Excel files with multiple sheets and extracting content
"""

import openpyxl
import pandas as pd
from pathlib import Path
import sys

def read_excel_file(file_path):
    """
    Read Excel file and extract content from all sheets
    """
    results = {
        'file_info': {},
        'sheets': {},
        'issues': []
    }
    
    try:
        # Check if file exists
        if not Path(file_path).exists():
            results['issues'].append(f"File not found: {file_path}")
            return results
            
        # Load workbook with openpyxl to get detailed info
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            results['file_info']['total_sheets'] = len(workbook.sheetnames)
            results['file_info']['sheet_names'] = workbook.sheetnames
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet_data = {
                    'name': sheet_name,
                    'content': [],
                    'images': [],
                    'dimensions': {},
                    'cell_count': 0
                }
                
                worksheet = workbook[sheet_name]
                
                # Get sheet dimensions
                if worksheet.max_row > 0 and worksheet.max_column > 0:
                    sheet_data['dimensions'] = {
                        'max_row': worksheet.max_row,
                        'max_column': worksheet.max_column
                    }
                    
                    # Extract cell content
                    for row in worksheet.iter_rows(values_only=True):
                        row_data = []
                        for cell in row:
                            if cell is not None:
                                row_data.append(str(cell))
                                sheet_data['cell_count'] += 1
                            else:
                                row_data.append('')
                        
                        # Only add non-empty rows
                        if any(cell for cell in row_data):
                            sheet_data['content'].append(row_data)
                
                # Check for images (basic detection)
                try:
                    if hasattr(worksheet, '_images') and worksheet._images:
                        sheet_data['images'] = [f"Image found at position {img.anchor}" for img in worksheet._images]
                    elif hasattr(worksheet, 'legacy_drawing') and worksheet.legacy_drawing:
                        sheet_data['images'].append("Legacy drawing/image detected")
                except Exception as e:
                    sheet_data['images'].append(f"Image detection error: {str(e)}")
                
                results['sheets'][sheet_name] = sheet_data
                
        except Exception as e:
            results['issues'].append(f"Error reading with openpyxl: {str(e)}")
            
            # Fallback to pandas
            try:
                excel_data = pd.read_excel(file_path, sheet_name=None)
                results['file_info']['total_sheets'] = len(excel_data.keys())
                results['file_info']['sheet_names'] = list(excel_data.keys())
                
                for sheet_name, df in excel_data.items():
                    sheet_data = {
                        'name': sheet_name,
                        'content': [],
                        'images': ['Cannot detect images with pandas fallback'],
                        'dimensions': {
                            'rows': len(df),
                            'columns': len(df.columns)
                        },
                        'cell_count': df.count().sum()
                    }
                    
                    # Convert DataFrame to list format
                    if not df.empty:
                        # Add headers
                        sheet_data['content'].append(df.columns.tolist())
                        # Add data rows
                        for _, row in df.iterrows():
                            sheet_data['content'].append([str(cell) if pd.notna(cell) else '' for cell in row])
                    
                    results['sheets'][sheet_name] = sheet_data
                    
            except Exception as e2:
                results['issues'].append(f"Error reading with pandas: {str(e2)}")
        
    except Exception as e:
        results['issues'].append(f"General error: {str(e)}")
    
    return results

def format_to_markdown(results, output_file):
    """
    Format Excel content to Markdown
    """
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("# Excel File Content\n\n")
        
        if results['issues']:
            f.write("## Issues Encountered\n\n")
            for issue in results['issues']:
                f.write(f"- {issue}\n")
            f.write("\n")
        
        if results['file_info']:
            f.write("## File Information\n\n")
            f.write(f"- Total Sheets: {results['file_info'].get('total_sheets', 'Unknown')}\n")
            f.write(f"- Sheet Names: {', '.join(results['file_info'].get('sheet_names', []))}\n\n")
        
        if results['sheets']:
            f.write("## Sheet Contents\n\n")
            
            for sheet_name, sheet_data in results['sheets'].items():
                f.write(f"### Sheet: {sheet_name}\n\n")
                
                if sheet_data['dimensions']:
                    f.write(f"**Dimensions:** {sheet_data['dimensions']}\n")
                f.write(f"**Cell Count:** {sheet_data['cell_count']}\n\n")
                
                if sheet_data['images']:
                    f.write("**Images/Graphics:**\n")
                    for img in sheet_data['images']:
                        f.write(f"- {img}\n")
                    f.write("\n")
                
                if sheet_data['content']:
                    f.write("**Content:**\n\n")
                    
                    # Create markdown table if data is structured
                    if len(sheet_data['content']) > 1:
                        # Assume first row is headers
                        headers = sheet_data['content'][0]
                        f.write("| " + " | ".join(headers) + " |\n")
                        f.write("| " + " | ".join(['---'] * len(headers)) + " |\n")
                        
                        for row in sheet_data['content'][1:]:
                            # Pad row to match header length
                            padded_row = row + [''] * (len(headers) - len(row))
                            f.write("| " + " | ".join(padded_row[:len(headers)]) + " |\n")
                    else:
                        # Single row, just display as list
                        for row in sheet_data['content']:
                            f.write("- " + ", ".join(row) + "\n")
                    
                    f.write("\n")
                else:
                    f.write("*No content found in this sheet*\n\n")

if __name__ == "__main__":
    excel_file = "Fruits.xlsx"
    output_file = "ExcelFile.md"
    
    print(f"Attempting to read Excel file: {excel_file}")
    results = read_excel_file(excel_file)
    
    print(f"Formatting results to: {output_file}")
    format_to_markdown(results, output_file)
    
    print("Process completed. Check Report.md for detailed results.")